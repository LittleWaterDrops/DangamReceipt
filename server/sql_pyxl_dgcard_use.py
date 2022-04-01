# -*- coding: utf-8 -*-
from turtle import st
import pymysql
import pprint
import pandas as pd

from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image

connect_sql = pymysql.connect(host='192.168.0.4', user='root', passwd='dg1q2w#E', charset='utf8')
cursor = connect_sql.cursor(pymysql.cursors.DictCursor) #pymysql.cursors.DictCursor

database_use = 'use DGcard_use_data_all;'
cursor.execute(database_use) #명령어 적용

select_table = 'select * from DGcard_use_data_2202;'
cursor.execute(select_table)

result = cursor.fetchall()
# print(type(result))
# pprint.pprint(result)

connect_sql.commit()
connect_sql.close()

df = pd.DataFrame(result)
df['NO'] = df['No']
df['증빙'] = '○'
df['일자'] = df['일자'].apply(str)

df1 = df[['NO', '일자', '구분', '사용처', '내용', '금액', '사용자', '비고', '증빙']]
money_sum = df['금액'].sum()
startday = df.at[0, '일자']
endday = df.at[df.shape[0]-1, '일자']


wb = Workbook()
path = '/home/receipt_automation/'
today = str(datetime.today())
today = today[:11]
month = datetime.today().month

#####################
### sheet : index ###
#####################
s_index = wb.active
s_index.title = 'index'
s_index.merge_cells('B2:C2')
s_index_img = Image(path + 'dg_LOGO_image/dangam_LOGO_margin.png')
s_index_img.height = 72
s_index_img.width = 194
s_index.add_image(s_index_img, 'B2')
# s_index.add_image(s_index_img, 'G2')

s_index.row_dimensions[1].height = 10
s_index.row_dimensions[2].height = 62.5
s_index.row_dimensions[3].height = 10
s_index.row_dimensions[4].height = 37
s_index.row_dimensions[5].height = 37
s_index.row_dimensions[6].height = 15
s_index.column_dimensions['A'].width = 3.375
s_index.column_dimensions['B'].width = 20
s_index.column_dimensions['C'].width = 20
s_index.column_dimensions['D'].width = 20

s_index.cell(2, 4).value = '기술 본부 \n하나카드 9436'
s_index.cell(2, 4).font = Font(size=14, bold=True, name='맑은 고딕')
s_index.cell(2, 4).alignment = Alignment(horizontal='center', vertical='center')

s_index.cell(4, 2).value = '사용기간'
s_index.cell(4, 2).font = Font(size=11, bold=True, name='맑은 고딕')
s_index.cell(4, 2).alignment = Alignment(horizontal='center', vertical='center')
s_index.cell(4, 2).fill = PatternFill(fgColor="FDEADA", fill_type="solid")
s_index.cell(4, 2).border = Border(top=Side(style='thick'), bottom=Side(style='thin'))
s_index.cell(4, 3).value = startday
s_index.cell(4, 3).font = Font(size=11, bold=True, name='맑은 고딕')
s_index.cell(4, 3).alignment = Alignment(horizontal='center', vertical='center')
s_index.cell(4, 3).border = Border(top=Side(style='thick'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
s_index.cell(4, 4).value = f'~{endday}'
s_index.cell(4, 4).font = Font(size=11, bold=True, name='맑은 고딕')
s_index.cell(4, 4).alignment = Alignment(horizontal='center', vertical='center')
s_index.cell(4, 4).border = Border(top=Side(style='thick'), bottom=Side(style='thin'))

s_index.merge_cells('C5:D5')
s_index.cell(5, 2).value = '총사용금액'
s_index.cell(5, 2).font = Font(size=11, bold=True, name='맑은 고딕')
s_index.cell(5, 2).alignment = Alignment(horizontal='center', vertical='center')
s_index.cell(5, 2).fill = PatternFill(fgColor="FDEADA", fill_type="solid")
s_index.cell(5, 2).border = Border(bottom=Side(style='thick'))
s_index.cell(5, 3).value = format(money_sum, ',') + ' '
s_index.cell(5, 3).font = Font(size=11, bold=True, name='맑은 고딕')
s_index.cell(5, 3).alignment = Alignment(horizontal='right', vertical='center')
s_index.cell(5, 3).border = Border(bottom=Side(style='thick'), left=Side(style='thin'))

# df_index = df[df['구분'].str.contains('식권 대체')]
# print(df_index)

# print(df_index['금액'].sum())

month_total_money = 0
month_total_count = 0
s_index_row7 = ['항목', '건수', '금액']
s_index_row7_list = ['공갈', '회식비', '식권 대체', '야근 식대', '소모품비', '복리후생비', '여비교통비', 
                     '통신비', '호스팅', '도서인쇄비', '접대비', '교육훈련비', '지급수수료', '회의비', "기타"]
for i in range(len(s_index_row7_list)):
    s_index.row_dimensions[i+7].height = 37
    for j in range(3):
        if i == 0:
            s_index.cell(i+7, j+2).value = s_index_row7[j]
            s_index.cell(i+7, j+2).font = Font(size=11, bold=True, name='맑은 고딕')
            s_index.cell(i+7, j+2).alignment = Alignment(horizontal='center', vertical='center')
            if j == 0:
                s_index.cell(i+7, j+2).border = Border(top=Side(style='thick'), left=Side(style='thick'))
            elif j == 1:
                s_index.cell(i+7, j+2).border = Border(top=Side(style='thick'), left=Side(style='thin'), right=Side(style='thin'))
            elif j == 2:
                s_index.cell(i+7, j+2).border = Border(top=Side(style='thick'), right=Side(style='thick'))
            s_index.cell(i+7, j+2).fill = PatternFill(fgColor="FDEADA", fill_type="solid")
        elif j == 0:
            s_index.cell(i+7, j+2).value = s_index_row7_list[i]
            s_index.cell(i+7, j+2).font = Font(size=11, bold=True, name='맑은 고딕')
            s_index.cell(i+7, j+2).alignment = Alignment(horizontal='center', vertical='center')
            s_index.cell(i+7, j+2).border = Border(left=Side(style='thick'))
            df_count = df[df['구분'].str.contains(s_index_row7_list[i])]
            df_count_money_sum = df_count['금액'].sum()
        elif j == 1:
            s_index.cell(i+7, j+2).value = f'{len(df_count.index)}건'
            s_index.cell(i+7, j+2).font = Font(size=11, name='맑은 고딕')
            s_index.cell(i+7, j+2).alignment = Alignment(horizontal='center', vertical='center')
            s_index.cell(i+7, j+2).border = Border(right=Side(style='thin'), left=Side(style='thin'))
            month_total_count += len(df_count.index)
        elif j == 2:
            s_index.cell(i+7, j+2).value = format(df_count_money_sum, ',') + ' '
            s_index.cell(i+7, j+2).font = Font(size=11, name='맑은 고딕')
            s_index.cell(i+7, j+2).alignment = Alignment(horizontal='right', vertical='center')
            s_index.cell(i+7, j+2).border = Border(right=Side(style='thick'))
            month_total_money += df_count_money_sum

    if i == len(s_index_row7_list) - 1:
        for k in range(3):
            s_index.row_dimensions[i+8].height = 37
            if k == 0:
                s_index.cell(i+8, k+2).value = '합계'
                s_index.cell(i+8, k+2).alignment = Alignment(horizontal='center', vertical='center')
                s_index.cell(i+8, k+2).border = Border(top=Side(style='thick'), bottom=Side(style='thick'),left=Side(style='thick'))
            elif k == 1:
                s_index.cell(i+8, k+2).value = f'{month_total_count}건'
                s_index.cell(i+8, k+2).alignment = Alignment(horizontal='center', vertical='center')
                s_index.cell(i+8, k+2).border = Border(top=Side(style='thick'), bottom=Side(style='thick'),left=Side(style='thin'), right=Side(style='thin'))
            elif k == 2:
                s_index.cell(i+8, k+2).value = format(month_total_money, ',') + ' '
                s_index.cell(i+8, k+2).alignment = Alignment(horizontal='right', vertical='center')
                s_index.cell(i+8, k+2).border = Border(top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'))

            s_index.cell(i+8, k+2).font = Font(size=11, name='맑은 고딕', bold=True)
            s_index.cell(i+8, k+2).fill = PatternFill(fgColor="FDEADA", fill_type="solid")



##########################
### sheet : 사용내역서 ###
##########################
use_statement = wb.active
use_statement = wb.create_sheet('사용내역서')
use_statement_img = Image(path + 'dg_LOGO_image/dgLOGO.png')
use_statement_img.height = 62
use_statement_img.width = 125
use_statement.add_image(use_statement_img, 'B2')

rows_num = [1, 2, 3, 4, 5, 6, 7]
rows_height = [10, 48, 10, 35, 35, 35, 35]
column_name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',]
column_width = [3.375, 5.625, 11.125, 15, 15, 5.5125, 5.5125, 5.5125, 5.5125, 5.5125, 6.625, 6.625, 11.25, 24.25, 5.5125]
for a in range(7):
    use_statement.row_dimensions[rows_num[a]].height = rows_height[a]

for b in range(15):
    use_statement.column_dimensions[column_name[b]].width = column_width[b]


use_statement.merge_cells('E2:M2')
use_statement.merge_cells('N2:O2')
use_statement.cell(2, 5).value = '법인카드 사용내역서'
use_statement.cell(2, 5).font = Font(size=20, bold=True, name='맑은 고딕')
use_statement.cell(2, 5).alignment = Alignment(horizontal='center', vertical='center')
use_statement.cell(2, 14).value = '하나카드 9436'
use_statement.cell(2, 14).font = Font(size=14, bold=True, name='맑은 고딕')
use_statement.cell(2, 14).alignment = Alignment(horizontal='center', vertical='center')


use_statement.merge_cells('B4:C4')
use_statement.merge_cells('D4:H4')
use_statement.merge_cells('I4:J4')
use_statement.merge_cells('K4:O4')
border_top_thick = Border(top=Side(style='thick'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
row4 = [2, 4, 9, 11]
row4_v = ['성명', '박윤도', '소속', '기술 본부']
for c in range(4):
    if row4[c] == 2:
        use_statement.cell(4, row4[c]).value = row4_v[c]
        use_statement.cell(4, row4[c]).font = Font(size=11, name='맑은 고딕')
        use_statement.cell(4, row4[c]).border = Border(top=Side(style='thick'), bottom=Side(style='thin'))
        use_statement.cell(4, row4[c]).alignment = Alignment(horizontal='center', vertical='center')
    else:
        use_statement.cell(4, row4[c]).value = row4_v[c]
        use_statement.cell(4, row4[c]).font = Font(size=11, name='맑은 고딕')
        use_statement.cell(4, row4[c]).border = Border(top=Side(style='thick'))
        use_statement.cell(4, row4[c]).alignment = Alignment(horizontal='center', vertical='center')
        use_statement.cell(4, row4[c]).border = border_top_thick


use_statement.merge_cells('B5:C5')
use_statement.merge_cells('D5:H5')
use_statement.merge_cells('I5:J5')
use_statement.merge_cells('K5:O5')
border_bottom_thick = Border(bottom=Side(style='thick'), right=Side(style='thin'), left=Side(style='thin'))
row5_v = ['직급', '사원', '작성일자', today]
row5 = [2, 4, 9, 11]
for d in range(4):
    if row5[d] == 2:
        use_statement.cell(5, row5[d]).value = row5_v[d]
        use_statement.cell(5, row5[d]).font = Font(size=11, name='맑은 고딕')
        use_statement.cell(5, row5[d]).alignment = Alignment(horizontal='center', vertical='center')
        use_statement.cell(5, row5[d]).border = Border(bottom=Side(style='thick'))
    else:
        use_statement.cell(5, row5[d]).value = row5_v[d]
        use_statement.cell(5, row5[d]).font = Font(size=11, name='맑은 고딕')
        use_statement.cell(5, row5[d]).alignment = Alignment(horizontal='center', vertical='center')
        use_statement.cell(5, row5[d]).border = border_bottom_thick

######### Data input form
use_statement.merge_cells('F7:J7')
use_statement.merge_cells('K7:L7')
border_row7 = Border(top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thin'), left=Side(style='thin'))
row7 = [2, 3, 4, 5, 6, 11, 13, 14, 15]
row7_v = ['NO.', '일자', '구분', '사용처', '내용', '금액', '사용자', '비고', '증빙']
for e in range(9):
    if row7[e] == 2 or row7[e] == 15:
        use_statement.cell(7, row7[e]).value = row7_v[e]
        use_statement.cell(7, row7[e]).font = Font(size=11, name='맑은 고딕')
        use_statement.cell(7, row7[e]).alignment = Alignment(horizontal='center', vertical='center')
        use_statement.cell(7, row7[e]).border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
    else:
        use_statement.cell(7, row7[e]).value = row7_v[e]
        use_statement.cell(7, row7[e]).font = Font(size=11, name='맑은 고딕')
        use_statement.cell(7, row7[e]).alignment = Alignment(horizontal='center', vertical='center')
        use_statement.cell(7, row7[e]).border = border_row7

######### Data 
border_row_data = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
row_input_data = [2, 3, 4, 5, 6, 11, 13, 14, 15]
row_data_column = ['NO', '일자', '구분', '사용처', '내용', '금액', '사용자', '비고', '증빙']
for f in range(df.shape[0]):
    use_statement.merge_cells(f'F{f+8}:J{f+8}')
    use_statement.merge_cells(f'K{f+8}:L{f+8}')
    use_statement.row_dimensions[f+8].height = 30
    for g in range(len(row_data_column)):
        if row_input_data[g] == 2:
            use_statement.cell(f+8, row_input_data[g]).value = df.at[f, row_data_column[g]]
            use_statement.cell(f+8, row_input_data[g]).font = Font(size=11, name='맑은 고딕')
            use_statement.cell(f+8, row_input_data[g]).alignment = Alignment(horizontal='center', vertical='center')
            use_statement.cell(f+8, row_input_data[g]).border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        elif row_input_data[g] == 15:
            use_statement.cell(f+8, row_input_data[g]).value = df.at[f, row_data_column[g]]
            use_statement.cell(f+8, row_input_data[g]).font = Font(size=11, name='맑은 고딕')
            use_statement.cell(f+8, row_input_data[g]).alignment = Alignment(horizontal='center', vertical='center')
            use_statement.cell(f+8, row_input_data[g]).border = Border(bottom=Side(style='thin'))
        else:
            use_statement.cell(f+8, row_input_data[g]).value = df.at[f, row_data_column[g]]
            use_statement.cell(f+8, row_input_data[g]).font = Font(size=11, name='맑은 고딕')
            if row_input_data[g] == 11:
                use_statement.cell(f+8, row_input_data[g]).value = df.at[f, row_data_column[g]]
                use_statement.cell(f+8, row_input_data[g]).alignment = Alignment(horizontal='right', vertical='center')
            elif row_input_data[g] == 14:
                use_statement.cell(f+8, row_input_data[g]).alignment = Alignment(horizontal='left', vertical='center')
            else:
                use_statement.cell(f+8, row_input_data[g]).alignment = Alignment(horizontal='center', vertical='center')
            use_statement.cell(f+8, row_input_data[g]).border = border_row_data

    if f == df.shape[0] - 1:
        use_statement.row_dimensions[f+9].height = 30
        use_statement.merge_cells(f'B{f+9}:J{f+9}')
        use_statement.merge_cells(f'K{f+9}:L{f+9}')
        use_statement.merge_cells(f'M{f+9}:O{f+9}')
        use_statement.cell(f+9, 2).value = '법인카드 소계(기술 본부 하나카드 9436)'
        use_statement.cell(f+9, 2).font = Font(size=11, name='맑은 고딕', bold=True)
        use_statement.cell(f+9, 2).alignment = Alignment(horizontal='center', vertical='center')
        use_statement.cell(f+9, 2).border = Border(top=Side(style='thin'), bottom=Side(style='thick'), right=Side(style='thin'))

        use_statement.cell(f+9, 11).value = format(money_sum, ',') + ' '
        use_statement.cell(f+9, 11).font = Font(size=11, name='맑은 고딕', bold=True)
        use_statement.cell(f+9, 11).alignment = Alignment(horizontal='right', vertical='center')
        use_statement.cell(f+9, 11).border = Border(top=Side(style='thin'), bottom=Side(style='thick'), left=Side(style='thin'))

        use_statement.cell(f+9, 13).border = Border(top=Side(style='thin'), bottom=Side(style='thick'), left=Side(style='thin'))
 

wb.save(path+f'{month - 1}월_기술본부_법인카드사용내역서.xlsx')
